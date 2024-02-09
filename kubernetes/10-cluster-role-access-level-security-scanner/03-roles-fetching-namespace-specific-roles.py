from kubernetes import client, config
import pandas as pd
from collections import defaultdict
import pytz
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def get_roles_and_bindings(api_instance, namespace="testtns"):
    if namespace:    
        roles = api_instance.list_namespaced_role(namespace=namespace)
    else:
        roles = api_instance.list_cluster_role()
    return roles.items

def extract_permissions(roles):
    # permissions = defaultdict(dict)
    permissions = defaultdict(lambda: defaultdict(lambda: {'verbs': set(), 'creation_timestamp': None, 'namespace': None}))

    
    for role in roles:
        # print(role.metadata)
        # print("====================================================")
        role_name = role.metadata.name
        if not role_name.startswith("system:") and role.metadata.namespace not in ["kube-node-lease", "kube-public", "kube-system"]:
            for rule in role.rules:
                if rule.resources:
                    for resource in rule.resources:
                        for verb in rule.verbs:
                            resource_prefix = resource.split("/")[0]
                            permissions[role_name][resource_prefix]['verbs'].add(verb)
                            creation_timestamp = role.metadata.creation_timestamp
                            permissions[role_name][resource_prefix]['creation_timestamp'] = creation_timestamp.strftime("%d-%b-%Y") if creation_timestamp else None
                            permissions[role_name][resource_prefix]['namespace'] = role.metadata.namespace
    # print(permissions)
    return permissions


def create_excel_sheet(permissions):
    flattened_permissions = []
    for role_name, resources in permissions.items():
        for resource_prefix, data in resources.items():
            verbs = ', '.join(sorted(data['verbs']))
            creation_timestamp = data['creation_timestamp']
            namespace = data['namespace']
            flattened_permissions.append((
                role_name,
                "Role",
                resource_prefix,
                ", ".join(sorted(data['verbs'])),
                creation_timestamp,
                namespace
            ))
    
    df = pd.DataFrame(flattened_permissions, columns=['Role/Binding Name', 'Type', 'Resource', 'Unique Verbs', 'Creation Timestamp', 'Namespace'])
    df.to_excel('kubernetes_permissions.xlsx', index=False)

     # Open the workbook
    wb = load_workbook('kubernetes_permissions.xlsx')
    ws = wb.active
    
    # Resize columns to fit content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save('kubernetes_permissions.xlsx')

def main():
    config.load_kube_config()
    v1 = client.RbacAuthorizationV1Api()

    roles = get_roles_and_bindings(v1)

    permissions = extract_permissions(roles)
    create_excel_sheet(permissions)

if __name__ == "__main__":
    main()
